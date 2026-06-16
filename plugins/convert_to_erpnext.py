"""
Convert SYS Daily Cost Sheet -> ERPNext Journal Entry Import Format

Logic per row:
  - Cost column (positive)  → Debit Src account, Credit Sub Cat account
  - Cost column (negative)  → Credit Src account, Debit Sub Cat account
  - Other column (positive) → Credit Src account, Debit Sub Cat account
  - Other column (negative) → Debit Src account, Credit Sub Cat account
  - Row with #N/A as Src    → flagged (SRC=N/A)
  - Row where Src == Sub Cat → flagged (SRC==SUB_CAT)
"""

import pandas as pd
from datetime import datetime
from pathlib import Path

COMPANY = "Seinn Yaung So MFG"
ENTRY_TYPE = "Journal Entry"
SERIES = "ACC-JV-.YYYY.-"
IS_OPENING = "No"


def extract_account_code(account_str):
    """Return the account string as-is (ERPNext uses full name)."""
    if pd.isna(account_str) or str(account_str).strip() in ("#N/A", ""):
        return None
    return str(account_str).strip()


def process_sheet(input_file, sheet_name):
    """Return (clean_rows, flagged_rows) for the given sheet."""
    df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)

    # Row 1 (index 1) is the header
    df.columns = df.iloc[1]
    df = df.iloc[2:].reset_index(drop=True)
    df = df.iloc[:, :8]
    df.columns = [
        "Date",
        "Place",
        "Src",
        "Category",
        "Sub Cat",
        "Particular",
        "Cost",
        "Other",
    ]

    clean_rows = []
    flagged_rows = []
    jv_counter = 0

    for _, row in df.iterrows():
        date = row["Date"]
        if pd.isna(date):
            continue

        # Convert serial date if needed
        if isinstance(date, (int, float)):
            try:
                date = pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(date))
            except Exception:
                continue
        elif not isinstance(date, (pd.Timestamp, datetime)):
            try:
                date = pd.Timestamp(date)
            except Exception:
                continue

        posting_date = date.strftime("%Y-%m-%d")

        src = extract_account_code(row["Src"])
        sub_cat = extract_account_code(row["Sub Cat"])
        particular = (
            str(row["Particular"]).strip() if not pd.isna(row["Particular"]) else ""
        )

        cost_raw = row["Cost"]
        other_raw = row["Other"]

        cost = float(cost_raw) if not pd.isna(cost_raw) else None
        other = float(other_raw) if not pd.isna(other_raw) else None

        if cost is not None and cost != 0:
            amount = cost
            use_cost = True
        elif other is not None and other != 0:
            amount = other
            use_cost = False
        else:
            continue

        src_is_na = (src is None) or (str(row["Src"]).strip() == "#N/A")

        abs_amount = abs(amount)

        if use_cost:
            if amount > 0:
                debit_acct = sub_cat
                credit_acct = src
            else:
                debit_acct = src
                credit_acct = sub_cat
        else:
            if amount > 0:
                debit_acct = sub_cat
                credit_acct = src
            else:
                debit_acct = src
                credit_acct = sub_cat

        # Detect flag reason
        flag_reason = None
        if src_is_na:
            flag_reason = "SRC=N/A"
            credit_acct = sub_cat
            debit_acct = sub_cat
        elif src is not None and sub_cat is not None and src == sub_cat:
            flag_reason = "SRC==SUB_CAT"

        note = (
            f"[{flag_reason} - needs manual account assignment] {particular}"
            if flag_reason
            else particular
        )

        jv_id = f"ACC-JV-2026-{jv_counter:05d}"
        jv_counter += 1

        pair = [
            {
                "ID": jv_id,
                "Company": COMPANY,
                "Entry Type": ENTRY_TYPE,
                "Series": SERIES,
                "Posting Date": posting_date,
                "Debit (Accounting Entries)": None,
                "Credit (Accounting Entries)": abs_amount,
                "Is Opening": IS_OPENING,
                "ID (Accounting Entries)": credit_acct,
                "Account (Accounting Entries)": credit_acct,
                "User Remark (Accounting Entries)": note,
            },
            {
                "ID": None,
                "Company": None,
                "Entry Type": None,
                "Series": None,
                "Posting Date": None,
                "Debit (Accounting Entries)": abs_amount,
                "Credit (Accounting Entries)": None,
                "Is Opening": None,
                "ID (Accounting Entries)": debit_acct,
                "Account (Accounting Entries)": debit_acct,
                "User Remark (Accounting Entries)": note,
            },
        ]

        if flag_reason:
            flagged_rows.extend(pair)
        else:
            clean_rows.extend(pair)

    return clean_rows, flagged_rows


def _write_xlsx(rows, output_file, header_color="1F4E79"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    columns = [
        "ID",
        "Company",
        "Entry Type",
        "Series",
        "Posting Date",
        "Debit (Accounting Entries)",
        "Credit (Accounting Entries)",
        "Is Opening",
        "ID (Accounting Entries)",
        "Account (Accounting Entries)",
        "User Remark (Accounting Entries)",
    ]

    out_df = pd.DataFrame(rows, columns=columns)

    wb = Workbook()
    ws = wb.active
    ws.title = "JV Import"

    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    for col_idx, h in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    jv_fill = PatternFill("solid", fgColor="DEEAF1")
    normal_font = Font(name="Arial", size=9)
    jv_font = Font(name="Arial", size=9, bold=True)

    for row_idx, (_, data_row) in enumerate(out_df.iterrows(), 2):
        is_jv_header = (
            not pd.isna(data_row["ID"]) if data_row["ID"] is not None else False
        )
        for col_idx, col in enumerate(columns, 1):
            val = data_row[col]
            if pd.isna(val) if not isinstance(val, str) else False:
                val = None
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = jv_font if is_jv_header else normal_font
            if is_jv_header:
                cell.fill = jv_fill

    col_widths = [22, 20, 15, 18, 14, 28, 28, 12, 45, 45, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    wb.save(output_file)


def run(input_file, output_file, sheets, output_flagged_file=None):
    """Process sheets and write output files. Returns (clean_path, flagged_path_or_None)."""
    all_clean = []
    all_flagged = []
    flagged_counts = {"SRC=N/A": 0, "SRC==SUB_CAT": 0}

    for sheet in sheets:
        try:
            clean, flagged = process_sheet(input_file, sheet)
            all_clean.extend(clean)
            all_flagged.extend(flagged)
            print(f"  {sheet}: {len(clean)//2} clean, {len(flagged)//2} flagged")
        except Exception as e:
            print(f"  Skipped {sheet}: {e}")

    if not all_clean and not all_flagged:
        print("No data processed.")
        return None, None

    # Count flagged by reason
    for row in all_flagged:
        remark = row.get("User Remark (Accounting Entries)") or ""
        if "SRC=N/A" in remark:
            flagged_counts["SRC=N/A"] += 1
        elif "SRC==SUB_CAT" in remark:
            flagged_counts["SRC==SUB_CAT"] += 1
    flagged_counts = {k: v // 2 for k, v in flagged_counts.items()}

    written_clean = None
    written_flagged = None

    if all_clean:
        _write_xlsx(all_clean, output_file)
        written_clean = output_file
        print(f"\nSaved (clean): {output_file}")
        print(f"  {len(all_clean)//2} journal entries")
    else:
        print("\nNo clean entries to write.")

    if all_flagged:
        if output_flagged_file is None:
            p = Path(output_file)
            output_flagged_file = str(p.with_stem(p.stem + "_flagged"))

        _write_xlsx(all_flagged, output_flagged_file, header_color="FF6B35")
        written_flagged = output_flagged_file
        total_flagged = len(all_flagged) // 2

        print(f"\nWARNING: {total_flagged} flagged journal entries need manual review")
        for reason, count in flagged_counts.items():
            if count:
                print(f"  - {reason:<14}: {count} entries")
        print(f"Saved (flagged): {output_flagged_file}")

    return written_clean, written_flagged


if __name__ == "__main__":
    run(
        input_file="/mnt/user-data/uploads/shh_Detail_Daily_Cost_-_Yangon__2_3_4_2026_.xlsx",
        output_file="/mnt/user-data/outputs/erpnext_jv_import.xlsx",
        sheets=[
            "RealReferenceSheet - Feb",
            "RealReferenceSheet - Mar",
            "RealReferenceSheet - Apr",
        ],
    )
